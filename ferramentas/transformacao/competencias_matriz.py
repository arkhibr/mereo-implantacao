"""
Skill: extração de competências a partir da matriz cargo×competência.

A fonte do cliente chega como uma MATRIZ (formato Belmont): na aba-título, cada
cargo ocupa um bloco de duas colunas — `Competências` e `Peso` — e as linhas
abaixo listam as competências daquele cargo com o peso (fração que soma 1). As
descrições de cada competência ficam em abas separadas, uma por cargo.

A plataforma RHTec/Mereo, porém, modela em dois níveis (ver linha de exemplo do
template do catálogo): `Competência` é a competência nomeada (ex. Comprometimento,
Tipo "Comportamental") e `Fator de Avaliação` é uma faceta abaixo dela.

Modelo adotado — **fator-espelho** (decidido com o usuário em 2026-06-09, com o
template como referência): cada competência da Belmont vira uma `Competência` da
plataforma com **um único fator** (espelho 1:1). O peso da matriz, que existe por
competência, cai limpo nesse fator. Os formulários (1 por cargo) referenciam
competência+fator por código e replicam o peso nos avaliadores em uso
(Gestor→LIDER, Autoavaliação→AUTO).

Posicionamento do conteúdo no catálogo (corrigido em 2026-06-23 após feedback do
time de implantação da Mereo — a competência é só o container, toda a substância
mora no fator, que é o item pontuado):
  - `Descrição da Competência` (col C): VAZIO.
  - `Nome do Fator de Avaliação` (col F): a DEFINIÇÃO da competência.
  - `Descrição do Fator de Avaliação` (col G): os COMPORTAMENTOS OBSERVÁVEIS
    (os níveis da escala), concatenados.
Nos formulários, os avaliadores não usados saem como 0 (não vazio): a plataforma
exige o preenchimento das 7 colunas para importar.

Esta ferramenta é pura: recebe o caminho do arquivo e devolve as duas tabelas já
no formato dos templates, mais o dicionário de códigos. Quem grava staging e lê
mapeamento são os agentes (`competencias`, `formularios`).
"""
import unicodedata

import openpyxl

# Colunas exatas dos templates (templates/competencias/).
COLUNAS_CATALOGO = [
    "Código da Competência", "Nome da competência", "Descrição da Competencia",
    "Tipo", "Código do Fator de Avaliação", "Nome do Fator de Avaliação",
    "Descrição do Fator de Avaliação",
]
COLUNAS_FORMULARIO = [
    "Código", "Descrição", "Código da Competência", "Código do Fator de Avaliação",
    "AUTO", "LIDER", "PAR", "TIME", "COMITÊ", "CLIENTE", "FORNECEDOR",
]
# Avaliadores em uso na fonte Belmont: Gestor→LIDER, Autoavaliação→AUTO.
# SUPOSIÇÃO A CONFIRMAR: o peso é da competência, replicado por avaliador (não
# há peso próprio por avaliador na fonte).
AVALIADORES_ATIVOS = ["AUTO", "LIDER"]
TIPO_PADRAO = "Comportamental"  # a fonte não informa o tipo; default do template.
# Separador dos comportamentos observáveis numa única célula (col G do catálogo).
# Quebra de linha espelha o layout empilhado da fonte (níveis da escala).
SEP_COMPORTAMENTOS = "\n"
_MARCADOR_COMPORTAMENTOS = "comportamentos observaveis"  # já normalizado (sem acento)


def _norm(s) -> str:
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return s.lower().strip()


def _ler_matriz(ws):
    """{cargo: [(competencia, peso), ...]} + {cargo: soma_dos_pesos}.

    Cargos vêm da linha 1 (cada nome de cargo está na coluna de `Competências`);
    o peso está na coluna imediatamente à direita; dados começam na linha 3.
    """
    rows = list(ws.iter_rows(min_row=1, max_row=60, max_col=40, values_only=True))
    if not rows:
        return {}, {}
    cargos = [(str(v).strip(), ci) for ci, v in enumerate(rows[0]) if v and str(v).strip()]
    matriz, somas = {}, {}
    for nome, ci in cargos:
        pares = []
        for r in rows[2:]:
            comp = r[ci] if ci < len(r) else None
            peso = r[ci + 1] if ci + 1 < len(r) else None
            txt = str(comp).strip() if comp else ""
            if not txt or txt.lower() == "soma":
                continue
            if isinstance(peso, (int, float)):
                pares.append((txt, float(peso)))
        if pares:
            matriz[nome] = pares
            somas[nome] = round(sum(p for _, p in pares), 4)
    return matriz, somas


_ABAS_NAO_CARGO = {"referencias", "resultado"}


def _ler_definicoes(wb, nomes_norm: set, aba_matriz: str) -> dict:
    """norm(competência) → {'definicao': str, 'comportamentos': [str, ...]}.

    O conteúdo vive nas abas de cargo (uma por cargo). Layout de cada competência:
        <Nome da competência>            | Avaliação
        <definição da competência>
        Comportamentos observáveis       |          | <peso>
        <comportamento nível 1>
        ...                              (4 níveis da escala)

    Varremos TODAS as abas de cargo — o nome do cargo na matriz nem sempre bate
    com o nome da aba (ex. 'Comprador e Controlador de Manutenção' vs aba
    'Comprador e Controlador Manut.') — e ficamos com a primeira ocorrência, já
    que a definição/comportamentos são iguais em qualquer cargo onde a competência
    apareça."""
    out = {}
    for aba in wb.sheetnames:
        if aba == aba_matriz or _norm(aba) in _ABAS_NAO_CARGO:
            continue
        cells = list(wb[aba].iter_rows(min_row=1, max_row=400, max_col=3, values_only=True))
        n = len(cells)
        for i in range(n):
            a = cells[i][0]
            if not a:
                continue
            na = _norm(a)
            if na in out or na not in nomes_norm:
                continue
            b = cells[i][1]
            # Definição: linha seguinte. Confirmamos a linha de competência pelo
            # marcador "Avaliação" na col B; na falta dele, por uma frase logo
            # abaixo (>20 chars) que não seja outro nome de competência.
            marcador = b and str(b).strip().lower() == "avaliação"
            prox = cells[i + 1][0] if i + 1 < n else None
            if marcador and prox:
                definicao = str(prox).strip()
            elif prox and len(str(prox).strip()) > 20 and _norm(prox) not in nomes_norm:
                definicao = str(prox).strip()
            else:
                definicao = ""
            out[na] = {"definicao": definicao,
                       "comportamentos": _coletar_comportamentos(cells, i + 1, nomes_norm)}
    return out


def _coletar_comportamentos(cells, inicio: int, nomes_norm: set) -> list:
    """A partir da linha `inicio`, acha o marcador 'Comportamentos observáveis' e
    coleta as linhas de texto seguintes (os níveis da escala) até o fim do bloco
    da competência (linha vazia, próximo 'Avaliação' ou próximo nome conhecido)."""
    n = len(cells)
    for j in range(inicio, n):
        aj, bj = cells[j][0], cells[j][1]
        if bj and str(bj).strip().lower() == "avaliação":
            break  # chegou na próxima competência sem achar o marcador
        if aj and _norm(aj) == _MARCADOR_COMPORTAMENTOS:
            comps = []
            for k in range(j + 1, n):
                ak, bk = cells[k][0], cells[k][1]
                if not ak or _norm(ak) in nomes_norm:
                    break
                if bk and str(bk).strip().lower() == "avaliação":
                    break
                comps.append(str(ak).strip())
            return comps
    return []


def extrair(caminho_xlsx: str, aba_matriz: str = None,
            tipo_padrao: str = TIPO_PADRAO) -> dict:
    """Extrai catálogo (competência×fator) e formulários da matriz do cliente.

    aba_matriz: nome da aba-título (a matriz). Se None, usa a primeira aba.

    Devolve dict com:
      catalogo     — list[dict] no formato COLUNAS_CATALOGO (1 linha/competência)
      formularios  — list[dict] no formato COLUNAS_FORMULARIO (1 linha/cargo×comp)
      dicionario   — list[dict] {nome, codigo_competencia, codigo_fator}
      somas        — {cargo: soma_pesos} (para o agente avisar se ≠ 1)
      avisos       — list[str]
    """
    wb = openpyxl.load_workbook(caminho_xlsx, read_only=True, data_only=True)
    nome_aba = aba_matriz or wb.sheetnames[0]
    matriz, somas = _ler_matriz(wb[nome_aba])

    # Competências únicas na ordem de aparição → códigos espelho CPT##/FT##.
    todas = []
    for pares in matriz.values():
        for c, _ in pares:
            if c not in todas:
                todas.append(c)

    defs = _ler_definicoes(wb, {_norm(c) for c in todas}, nome_aba)
    wb.close()

    avisos = []
    cod = {c: (f"CPT{i + 1:02d}", f"FT{i + 1:02d}") for i, c in enumerate(todas)}

    catalogo, dicionario = [], []
    sem_desc, sem_comp = [], []
    for c in todas:
        cpt, ft = cod[c]
        info = defs.get(_norm(c), {})
        definicao = info.get("definicao", "")
        comportamentos = info.get("comportamentos", [])
        if not definicao:
            sem_desc.append(c)
        if not comportamentos:
            sem_comp.append(c)
        catalogo.append({
            "Código da Competência": cpt, "Nome da competência": c,
            "Descrição da Competencia": "",  # vazio: substância mora no fator
            "Tipo": tipo_padrao,
            "Código do Fator de Avaliação": ft,
            "Nome do Fator de Avaliação": definicao,  # definição da competência
            "Descrição do Fator de Avaliação": SEP_COMPORTAMENTOS.join(comportamentos),
        })
        dicionario.append({"nome": c, "codigo_competencia": cpt, "codigo_fator": ft})

    formularios = []
    for cargo, pares in matriz.items():
        cod_form = "FORM-" + _norm(cargo).upper().replace(" ", "-")[:24]
        for c, peso in pares:
            cpt, ft = cod[c]
            p = round(peso * 100)
            linha = {"Código": cod_form, "Descrição": cargo,
                     "Código da Competência": cpt, "Código do Fator de Avaliação": ft}
            for av in COLUNAS_FORMULARIO[4:]:
                # Avaliador não usado vai 0 (não vazio): a plataforma exige as 7
                # colunas preenchidas para importar.
                linha[av] = p if av in AVALIADORES_ATIVOS else 0
            formularios.append(linha)

    if sem_desc:
        avisos.append(f"{len(sem_desc)} competência(s) sem definição na fonte: {sem_desc}")
    if sem_comp:
        avisos.append(f"{len(sem_comp)} competência(s) sem comportamentos observáveis na fonte: {sem_comp}")
    ruins = {c: s for c, s in somas.items() if abs(s - 1.0) > 0.001}
    if ruins:
        avisos.append(f"Soma de pesos ≠ 1 em: {ruins}")

    return {"catalogo": catalogo, "formularios": formularios,
            "dicionario": dicionario, "somas": somas, "avisos": avisos}
